import openpyxl
exceldoc = openpyxl.load_workbook('excel2.xlsx')
sheet1 = exceldoc.active#Active Sheet
#excelcols.txt A,B,C,D,.......,Z,AA,AB,...CD
#Read the file
cols = ['A','B','C','D','E','F','G','H']
max_cell = str(cols[sheet1.max_column-1])+str(sheet1.max_row)
cells = sheet1['A1':max_cell]#Capture the last cell with data

for row in cells:
	for col in row:
		print(str(col.value) + '\t', end='')
	print('\n', end='')

exceldoc.close()